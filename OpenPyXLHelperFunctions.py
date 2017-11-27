import os
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_interval
import openpyxl
import re

import pandas as pd
import numpy as np

#gets data from a range
#sample execution
	#wb = load_workbook('AVV045751.xlsx',read_only = True, data_only = True)

	#strips = data_from_range('C9:J442','Strips',wb)

# wb = load_workbook('Sample.xlsx', data_only = True)
# ws = wb.get_sheet_by_name('Data')
# rng = 'B4:C7'

def clean_range_name(rngString):
    return(re.sub("'","",rngString))

def data_from_range(RngAsString,WsAsString,wb):

    RngAsString = clean_range_name(RngAsString)
    WsAsString = clean_range_name(WsAsString)
    
    ws = wb.get_sheet_by_name(WsAsString)
    rng = RngAsString
    
    try:
        start,end = rng.split(':')
    except:
        start = rng
        end = start
#    StartCol,EndCol = Get_Ranges(rng)
    
    data = []
    
    for row in ws[start:end]:
        data.append([cell.value for cell in row])
        
    df = pd.DataFrame(data)
    
    return(df)


 
def data_to_range(DataAsDataframe,FirstCell,WsAsString,wb):

    WsAsString = clean_range_name(WsAsString)

    
    ws = wb.get_sheet_by_name(WsAsString)
    data = DataAsDataframe
    rng = FirstCell
    
    col =  "".join(re.findall("[A-Z]",rng, flags = re.I))
        
    col = col_letter_to_number(col)
    row = int("".join(re.findall("[0-9]",rng)))      
        
    try:
        NumRows,NumCols = data.shape           
        
        Rows = np.arange(row, row + NumRows,1)
        Cols = np.arange(col, col + NumCols,1)
        
    
        for XIndex,x in enumerate(Cols):
            for YIndex,y in enumerate(Rows):           
                ws.cell(row = y, column = x).value = data.iloc[YIndex,XIndex]
    except:
        ws.cell(row = row, column = col).value = data
        
    
            


def Clear_Range(RangeAsString,ws,wb):
    for row in ws[RangeAsString]:
        for cell in row:
            cell.value = None

    


####################################
#Helper Functions
###################################            
def Get_Ranges(RangeAsString):
    rng = RangeAsString
    
    
    #in case there is only one range being given
    try:
        start,end = rng.split(':')
    except:
        start = rng
        end = start   

        
    StartCol =  "".join(re.findall("[A-Z]",start, flags = re.I))
    EndCol =  "".join(re.findall("[A-Z]",end, flags = re.I))
    
    return[StartCol,EndCol]            

def col_letter_to_number(ColLetter):
    letter = ColLetter.lower()
    
    alphabet = list('abcdefghijklmnopqrstuvwxyz')
    
    
    if len(letter) == 1:
        ColNum = alphabet.index(letter)+1
                               
        return(ColNum)
                        
    else:
        Set = letter[0]
        SetNum = alphabet.index(Set) + 1
        SetNum = SetNum * 26
        
        Letter = letter[1]
        
        
        ColNum = alphabet.index(Letter) + 1
                               
        return(SetNum + ColNum)
    
